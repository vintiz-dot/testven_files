import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook and select active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Literature Review"

# Enable grid lines explicitly
ws.views.sheetView[0].showGridLines = True

# Data definition
headers = [
    "Candidate's Name",
    "Article Title",
    "Key Themes",
    "Purpose of Article",
    "Relevant Data",
    "Data Reliability and Validity",
    "Usefulness/Implementation Relevance",
    "Link"
]

data = [
    [
        "victor",
        "Quality EdTech Professional Development for K-12 Classroom Practice",
        "The disconnect between PD quantity and integration quality",
        "Investigates relationship between EdTech PD and classroom application",
        "Correlation between PD hours and collaborative design capability: r = -0.276 (p<0.018). Correlation with active modeling: r = 0.377 (p=0.001)",
        "IRB-approved quantitative survey (n=73) with IP blocking controls and statistical correlation analysis",
        "Demonstrates that traditional PD fails without active modeling (coaching)",
        "https://www.jdet.net/download/quality-edtech-professional-development-for-k12-classroom-practice-14809.pdf"
    ],
    [
        "victor",
        "The Effects of Coaching Teachers to Integrate Technology",
        "Coaching impact on student critical thinking",
        "Examines fifth-grade teacher coaching and student outcomes",
        "Coached group (n=64) showed higher student performance and critical thinking behaviors than control group (n=68)",
        "Experimental mixed-methods design with controlled and intervention groups",
        "Provides direct evidence that coaching increases technology purposefulness and student performance",
        "https://www.ilkogretim-online.org/index.php/pub/article/view/7400"
    ],
    [
        "victor",
        "Active Learning Statistics 2026",
        "Active vs. passive learning engagement",
        "Measures engagement differences across instructional models",
        "Active sessions yield 13x more learner talk time and 62.7% participation rate vs. 5% in lectures",
        "Large-scale platform analytics and empirical analysis by Engageli (2024)",
        "Quantifies the exact engagement ROI of moving from passive to active learning paradigms",
        "https://www.engageli.com/blog/active-learning-statistics-2026"
    ],
    [
        "victor",
        "Active vs Passive Learning with Technology",
        "VR and active retention",
        "Compares retention in VR simulations vs passive environments",
        "Learners actively engaging in VR retained 62% of task steps compared to 51% in passive settings",
        "Experimental cohort studies with 25-year longitudinal researcher perspective",
        "Validates the cognitive retention benefits of active technology creation and simulation",
        "https://transfrinc.com/resources/blog/active-learning-vs-passive-learning-with-technology"
    ],
    [
        "victor",
        "Exploring The Role of Instructional Coaching in A Shifting Classroom Technology Landscape",
        "Efficacy of the BDA coaching cycle",
        "Explores how coaching behaviors affect integration",
        "Formal 1:1 coaching sessions most significantly impact integration practice over general PL sessions",
        "Peer-reviewed qualitative and quantitative analysis of coaching frameworks",
        "Proves that 1:1 job-embedded coaching is superior to sit-and-get PD for integration",
        "https://files.eric.ed.gov/fulltext/EJ1472024.pdf"
    ],
    [
        "victor",
        "Digital Storytelling in Economics Subjects",
        "Active content creation and achievement",
        "Tests digital storytelling effectiveness",
        "Experimental group post-test mean: 75.21 (p=0.001) vs Control group: 54.97",
        "Randomized controlled pedagogical experiment across six academies (n=856)",
        "Strong statistical proof that student-generated digital content directly improves test scores",
        "https://files.eric.ed.gov/fulltext/EJ1407229.pdf"
    ],
    [
        "victor",
        "Enhancing Classroom Learning: The Impact of AI-Based Instructional Strategies",
        "Technology integration and student engagement",
        "Highlights the necessity of ongoing PD for AI integration",
        "Teachers lacking integration training struggle with engagement; AI strategies significantly enhance outcomes",
        "Mixed-methods literature synthesis and observational study",
        "Highlights the need for coaches to help teachers navigate emerging AI tools effectively",
        "https://rsisinternational.org/journals/ijriss/articles/enhancing-classroom-learning-the-impact-of-ai-based-instructional-strategies-on-student-engagement-and-outcomes/"
    ],
    [
        "victor",
        "Active learning increases student performance in science engineering and mathematics",
        "Active learning reducing achievement gaps",
        "Meta-analysis of active vs passive classrooms",
        "Reforming STEM to active learning significantly reduces achievement gaps and increases retention",
        "Meta-analysis of 1294 coded sources and 41 disaggregated studies published in PNAS",
        "Macro-level validation that active instructional models are necessary for academic equity",
        "https://www.pnas.org/doi/10.1073/pnas.1916903117"
    ],
    [
        "victor",
        "A Systematic Review of Pre-service Teachers' Digital Competence",
        "TPACK and digital creation weaknesses",
        "Analyzes competence evolution (2014-2024)",
        "Teachers show adequate daily digital skills but particular weaknesses in digital content creation and pedagogical integration",
        "Systematic review of 38 studies from Web of Science and Scopus",
        "Identifies the exact pedagogical gap (content creation) that an EdTech coach must fill",
        "https://www.frontiersin.org/journals/psychology/articles/10.3389/fpsyg.2025.1710983/full"
    ],
    [
        "victor",
        "The Impact of Teacher Professional Development Programs Incorporating Educational Technology",
        "EdTech PD and student achievement",
        "Ascertains the role of tech PD on student outcomes",
        "Identifies significant rises in student achievement when tech PD is properly contextualized and supported",
        "Comprehensive meta-analysis of recent educational technology integrations",
        "Supports the business case that high-quality tech PD directly links to student success",
        "https://jurnal.ucy.ac.id/index.php/fkip/article/view/2544"
    ],
    [
        "victor",
        "Meta-Analysis of Using Technology in Mathematics Education",
        "Technology impact on higher math",
        "Reviews digital technologies enhancing learning outcomes",
        "Strong positive correlation between technology's professional relevance and student engagement; 88 studies reviewed",
        "Extensive meta-analysis of 88 peer-reviewed studies",
        "Demonstrates the necessity of purposeful tool selection which coaches can facilitate",
        "https://www.mdpi.com/2227-7102/15/11/1544"
    ],
    [
        "victor",
        "The Strategic Paradigm of Purposeful Screen Time in K-12 Education",
        "Purposeful vs passive screen time",
        "Examines policy shifts toward active tech use",
        "Major districts are shifting policy to structured time-bound active participation rather than passive usage limits",
        "Policy analysis and contemporary research synthesis (2024-2026)",
        "Aligns the coaching program with contemporary district leadership trends and policies",
        "https://www.letsgolearn.com/math-assessment/the-strategic-paradigm-of-purposeful-screen-time-in-k-12-education-2024-2026-comprehensive-analysis/"
    ],
    [
        "victor",
        "Fostering Powerful Use of Technology Through Instructional Coaching",
        "Dynamic Learning Project outcomes",
        "Explores coaching's potential to close the use divide",
        "Sustained onsite coaching across 50 schools moved tech use from drill-and-practice to critical thinking",
        "Multi-state pilot project observation and longitudinal programmatic assessment",
        "Provides a scalable blueprint for deploying a school-wide coaching program successfully",
        "https://digitalpromise.org/2018/08/13/fostering-powerful-use-of-technology-through-instructional-coaching/"
    ],
    [
        "victor",
        "Digital Learning and Professional Learning Communities",
        "PLCs and digital innovation",
        "Examines how collaborative frameworks improve outcomes",
        "PLCs expand digital competencies rapidly allowing educators to adapt innovative practices over traditional PD",
        "Review of collaborative frameworks and digital integration",
        "Shows how coaches can leverage PLCs to scale active tech integration school-wide",
        "https://pmc.ncbi.nlm.nih.gov/articles/PMC12303314/"
    ],
    [
        "victor",
        "The Impact of Student-Centered Learning on Academic Motivation",
        "Autonomy and tech integration",
        "Analyzes personalization and active learning",
        "Empowering students to manage learning via multimedia creation increases engagement and achievement",
        "Comparative research between traditional instruction and student-centered approaches",
        "Reinforces the pedagogical core of shifting students from consumers to creators",
        "https://www.researchgate.net/publication/376973509_The_Impact_of_Student-Centered_Learning_on_Academic_Motivation_and_Achievement_A_Comparative_Research_between_Traditional_Instruction_and_Student-Centered_Approach"
    ]
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header

# Write data
for row_num, row_data in enumerate(data, 2):
    for col_num, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = val

# Premium Formatting styles
# Color palette: Dark charcoal / Slate navy header with white text
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2B3A42", end_color="2B3A42", fill_type="solid") # Dark slate navy

body_font = Font(name="Segoe UI", size=10)
link_font = Font(name="Segoe UI", size=10, color="1B65A6", underline="single") # Professional steel blue link

thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)

# Apply header formatting
ws.row_dimensions[1].height = 28
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Apply body formatting
for row_num in range(2, len(data) + 2):
    ws.row_dimensions[row_num].height = 55 # Generous height for readability with word wrap
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = thin_border
        
        # Link column style
        if col_num == 8: # Link is column 8 (H)
            cell.font = link_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            # Create actual hyperlink in Excel
            if cell.value:
                cell.hyperlink = cell.value
        else:
            cell.font = body_font
            if col_num == 1: # Name
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Set custom column widths
col_widths = {
    "A": 18,  # Candidate Name
    "B": 28,  # Article Title
    "C": 22,  # Key Themes
    "D": 28,  # Purpose
    "E": 35,  # Relevant Data
    "F": 30,  # Data Reliability
    "G": 30,  # Usefulness
    "H": 25   # Link
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save the workbook
output_path = "Victor_Moronu_Literature_Review.xlsx"
wb.save(output_path)
print(f"Excel file created successfully: {os.path.abspath(output_path)}")
